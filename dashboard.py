# Contents of ~/my_app/streamlit_app.py
import streamlit as st
import pandas as pd
import plotly.express as px
from openpyxl import load_workbook
from datetime import datetime
import plotly.graph_objs as go
from plotly.graph_objs import *
import numpy as np

spreadsheet_name = "RCSS Data Structure.xlsx"

st.title("Dashboard Reports - Rural Sentiments Survey (RSS)")

def display_df(tab_sheetname,skip_row):
    df = pd.read_excel(spreadsheet_name, sheet_name=tab_sheetname,skiprows=skip_row)
    return df

def main_page():
    try:
        df = pd.read_excel(spreadsheet_name, sheet_name="table1")
        df['Target Achieved'] = df['Target Achieved'].round()
        df['Pending'] = df['Pending'].round()
        wb = load_workbook(spreadsheet_name, data_only=True)
        sh = wb["table1_header"]
        from_date = str(sh["b2"].value)
        from_date = datetime.strptime(from_date, "%Y-%m-%d %H:%M:%S").strftime("%d-%B-%Y")
        to_date = str(sh["b3"].value)
        to_date = datetime.strptime(to_date, "%Y-%m-%d %H:%M:%S").strftime("%d-%B-%Y")
        fig = px.bar(df, x="Centre", y=["Target Achieved","Pending"],
                     text_auto=True
                     )
        # Set y-axes,x-axes titles
        fig.update_yaxes(title_text="<b>in percent</b>", secondary_y=False)
        fig.update_xaxes(title_text="<b>center</b>")
        fig.update_layout(title=dict(text='<b>Target Achieved\n(From: </b>' + from_date + ' <b>To: </b>' + to_date + ")",
        x=0.5,
        y=0.95,
        font=dict(
            family="Arial",
            size=20,
            color='#000000'
        )))
        fig.update_layout(legend_title="")
        c1,c2 = st.columns(2)
        with c1:
            st.dataframe(display_df("table1",0))
        with c2:
            st.plotly_chart(fig)
    except Exception as err:
        st.error(err)


def page_2_col1_common(table_sheetname,col_0,col_4,title,color):
    try:
        df = pd.read_excel(spreadsheet_name, sheet_name=table_sheetname,skiprows=1)
        df = df[[col_0,col_4]]
        df[[col_4]] = df[[col_4]].round(2)
        #
        # df[col_1] = df[col_1].round(decimals=2)
        # df[col_2] = df[col_2].round(2)
        # df[col_3] = df[col_3].round(2)
        # df[col_4] = df[col_4].round(2)
        # df['Pending'] = df['Pending'].round()
        # # wb = load_workbook(spreadsheet_name, data_only=True)
        # # sh = wb["table1_header"]
        # # from_date = str(sh["b2"].value)
        # # from_date = datetime.strptime(from_date, "%Y-%m-%d %H:%M:%S").strftime("%d-%B-%Y")
        # # to_date = str(sh["b3"].value)
        # # to_date = datetime.strptime(to_date, "%Y-%m-%d %H:%M:%S").strftime("%d-%B-%Y")
        # fig = px.histogram(df, x=col_0
        #              )
        # # Set y-axes,x-axes titles
        # fig.update_yaxes(title_text="", secondary_y=False)
        # fig.update_xaxes(title_text="")
        # fig.update_layout(title=dict(text="<b>" + title + "</b>",
        # x=0.5,
        # y=0.95,
        # font=dict(
        #     family="Arial",
        #     size=20,
        #     color='#000000'
        # )))
        # fig.update_layout(legend_title="")
        # return fig
        fig = go.Figure()
        fig.add_trace(
            go.Bar(x=df[col_0],
                   y=df[col_4],
                   marker=dict(
                       color=color,)
                   ))
        fig.update_layout(barmode='stack')
        fig.update_layout(title=dict(text='<b>' + title + '</b>'))
        return fig
    except Exception as err:
        st.error(err)


def page_2_col2_common(table_sheetname,col_0,col_1,col_2,col_3,col_4,title):
    try:
        df = pd.read_excel(spreadsheet_name, sheet_name=table_sheetname,skiprows=1)
        df[[col_1, col_2, col_3, col_4]] = df[[col_1, col_2, col_3, col_4]].round(2)
        #
        # df[col_1] = df[col_1].round(decimals=2)
        # df[col_2] = df[col_2].round(2)
        # df[col_3] = df[col_3].round(2)
        # df[col_4] = df[col_4].round(2)
        # df['Pending'] = df['Pending'].round()
        # # wb = load_workbook(spreadsheet_name, data_only=True)
        # # sh = wb["table1_header"]
        # # from_date = str(sh["b2"].value)
        # # from_date = datetime.strptime(from_date, "%Y-%m-%d %H:%M:%S").strftime("%d-%B-%Y")
        # # to_date = str(sh["b3"].value)
        # # to_date = datetime.strptime(to_date, "%Y-%m-%d %H:%M:%S").strftime("%d-%B-%Y")
        fig = px.bar(df, x=col_0, y=[col_1, col_2, col_3],
                     )
        # Set y-axes,x-axes titles
        fig.update_yaxes(title_text="", secondary_y=False)
        fig.update_xaxes(title_text="")
        fig.update_layout(title=dict(text="<b>" + title + "</b>",
        x=0.5,
        y=0.95,
        font=dict(
            family="Arial",
            size=20,
            color='#000000'
        )))
        fig.update_layout(legend_title="")
        return fig
        # print(df.dtypes)
    except Exception as err:
        st.error(err)

def page2():
    try:
        c01,c02 = st.columns(2)
        c1,buffer,c2 = st.columns(3)
        c3,buffer,c4 = st.columns(3)
        with c01:
            st.dataframe(display_df("table2_1",1))
        with c02:
            st.dataframe(display_df("table2_2",1))
        with c1:
            st.plotly_chart(page_2_col1_common("table2_1", "Centers", "Net Responses","Current Economic Conditions\n(Net Responses)","blue"))
        with c2:
            st.plotly_chart(page_2_col1_common("table2_2", "Centers", "Net Responses","Future Economic Conditions\n(Net Responses)","green"))
        with c3:
            st.plotly_chart(page_2_col2_common("table2_1", "Centers", "Improve", "Remains same", "Worsen", "Net Responses","Current Economic Conditions"))
        with c4:
            st.plotly_chart(page_2_col2_common("table2_2", "Centers", "Improve", "Remains same", "Worsen", "Net Responses","Future Economic Conditions"))
    except Exception as err:
        st.error(err)



def page4_common(table_sheetname,label_col_name,proportion_col_name,donut_title):
    try:
        df = pd.read_excel(spreadsheet_name, sheet_name=table_sheetname)
        wb = load_workbook(spreadsheet_name, data_only=True)
        sh = wb["table4_header"]
        from_date = str(sh["b2"].value)
        from_date = datetime.strptime(from_date, "%Y-%m-%d %H:%M:%S").strftime("%d-%B-%Y")
        to_date = str(sh["b3"].value)
        to_date = datetime.strptime(to_date, "%Y-%m-%d %H:%M:%S").strftime("%d-%B-%Y")
        data = {
            "hole": 0.8,
            "type": "pie",
            "labels": df[label_col_name],
            "values": df[proportion_col_name],
            "showlegend": True
        }
        data = (data)
        layout = {
            "title": donut_title,
            "font" : dict(
            family="Arial",
            size=15,
            color='#000000'
        )
        }
        fig = Figure(data=data, layout=layout)
        fig.update_traces(textinfo='value')
        return fig
    except Exception as err:
        st.error(err)



def page4():
    try:
        c01,c02,c03,c04,c05 = st.columns(5)
        with c01:
            st.dataframe(display_df("table4_1",0))
        with c02:
            st.dataframe(display_df("table4_2",0))
        with c03:
            st.dataframe(display_df("table4_3",0))
        with c04:
            st.dataframe(display_df("table4_4",0))
        with c05:
            st.dataframe(display_df("table4_5",0))
        c1,buffer,c2 = st.columns(3)
        c3,buffer,c4 = st.columns(3)
        with c1:
            st.plotly_chart(page4_common("table4_1","Type of Respondent - Gender Wise","Proportion","Gender wise"))
        with c2:
            st.plotly_chart(page4_common("table4_2","Type of Respondent - Occupation Wise","Proportion","Occupation Wise"))
        with c3:
            st.plotly_chart(page4_common("table4_3","Type of Respondent - Age Wise","Proportion","Age Wise"))
        with c4:
            st.plotly_chart(page4_common("table4_4","Type of Respondent - Education Wise","Proportion","Education Wise"))
        st.plotly_chart(page4_common("table4_5","Type of Respondent - Monthly Average Income Wise","Proportion","Monthly Average Income Wise"))

    except Exception as err:
        st.error(err)


page_names_to_funcs = {
    "Status of Completion of Field Work": main_page,
    "Parameter-wise Perceptions and Expectations":page2,
    "Profile of Respondents ": page4

}

selected_page = st.sidebar.selectbox("Select Report Topic", page_names_to_funcs.keys())
page_names_to_funcs[selected_page]()
